#!/usr/bin/env python3
"""
Cardiotrack Gmail Sync & Dashboard Builder
==========================================
Runs daily at 00:00. Does three things:
  1. Checks Gmail for the 3 report emails → downloads attachments to Data/Input/
  2. Reads all Excel files (including Quarter Target.xlsx) → builds dashboard_data.json
  3. Injects fresh JSON into Cardiotrack_Dashboard.html

Usage:
  python3 cardiotrack_sync.py            # full sync + rebuild
  python3 cardiotrack_sync.py --rebuild  # skip Gmail, just rebuild from existing Excel
"""

import os, sys, json, base64, hashlib, logging, csv, re
from datetime import datetime, timezone
from pathlib import Path

# ── pip install google-auth-oauthlib google-api-python-client openpyxl pandas ──
try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    GMAIL_AVAILABLE = True
except ImportError:
    GMAIL_AVAILABLE = False
    print("⚠  Gmail packages not installed. Run:  pip3 install google-auth-oauthlib google-api-python-client")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠  Excel package not installed. Run:  pip3 install openpyxl")

# ─── PATHS ────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent.parent          # …/Cardiotrack sales/
INPUT  = BASE / "Data" / "Input"
OUTPUT = BASE / "Data" / "Output"
AUTO   = BASE / "Automation"

QUARTER_TARGET_FILE = INPUT  / "Quarter Target.xlsx"
# The 3 Gmail-downloaded reports are CSV (xlsx attachments had inconsistent
# layouts — empty-month fill-down was being silently dropped on read).
INCOMING_FILE       = INPUT  / "Incoming_Order_Count_Insurer_Wise.csv"
REVENUE_FILE        = INPUT  / "Revenue_Generated_Insurer_Wise.csv"
CLOSED_FILE         = INPUT  / "Closed_Case_Count_Insurer_Wise.csv"
BILLING_FILE        = INPUT  / "Daily_Insurer_Billing_Data.xls"
# Detailed transaction-level billing files (category breakdown: package, ancillary,
# videography, service charges, home visit, interpretation).  Two files because the
# Zoho report is split when row-count exceeds the export limit.
DETAILED_BILLING_FILE1 = INPUT / "Daily_Insurer_Billing.xls"
DETAILED_BILLING_FILE2 = INPUT / "Daily_Insurer_Billing_2.xls"

DASHBOARD_HTML = OUTPUT / "Cardiotrack_Dashboard.html"
DATA_JSON      = OUTPUT / "dashboard_data.json"
SYNC_LOG       = AUTO   / "sync_log.json"
CREDS_FILE     = AUTO   / "credentials.json"
TOKEN_FILE     = AUTO   / "token.json"

# Gmail for the daily report attachments + Drive so the cloud-hosted GitHub
# Action can pull Quarter Target.xlsx from your Drive (auto-synced from Mac).
# Local runs don't need Drive — the script reads Quarter Target from the local
# Data/Input folder regardless — but having Drive in the scope keeps one token
# usable in both modes.
SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# ─── Email subjects → target filenames ───────────────────────────────────────
# The real emails arrive with subject:  Check out the "<report name>" report
# The Gmail search query matches the inner quoted string; we keep the exact
# subject form here so we can also tighten matching after the API returns hits.
# CSV reports (Zoho sends these as .csv attachments)
CSV_EMAIL_MAP = {
    'Incoming Order Count Insurer Wise': INCOMING_FILE,
    'Revenue Generated Insurer Wise':    REVENUE_FILE,
    'Closed Case Count Insurer Wise':    CLOSED_FILE,
}

# Excel/XLS reports (Zoho sends these as .xls / .xlsx attachments)
XLS_EMAIL_MAP = {
    'Daily Insurer Billing Data':        BILLING_FILE,
    # Detailed category-level billing exports.
    # ⚠  BOTH Daily_Insurer_Billing.xls and Daily_Insurer_Billing_2.xls are sent
    #    in the SAME email thread with subject "Daily Insurer Billing" (two separate
    #    messages, same subject).  Only DETAILED_BILLING_FILE1 is listed here so
    #    the main sync loop downloads it normally.  DETAILED_BILLING_FILE2 is
    #    handled by _sync_billing_file2() which scans the same thread by filename.
    'Daily Insurer Billing':             DETAILED_BILLING_FILE1,
}

# Attachment filename → destination for files that share a subject with another
# report and can't be matched by subject alone.
BILLING_EXTRA_FILES: dict = {
    'daily_insurer_billing_2.xls': DETAILED_BILLING_FILE2,
}

# Combined map (used by dashboard to know all expected reports)
EMAIL_MAP = {**CSV_EMAIL_MAP, **XLS_EMAIL_MAP}

# Full subject prefix/suffix on incoming Gmail messages.
SUBJECT_PREFIX = 'Check out the "'
SUBJECT_SUFFIX = '" report'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(AUTO / "sync.log", encoding="utf-8"),
    ]
)
log = logging.getLogger("cardiotrack")


# ══════════════════════════════════════════════════════════════════════════════
# 1.  GMAIL SYNC
# ══════════════════════════════════════════════════════════════════════════════

def load_sync_log() -> dict:
    if SYNC_LOG.exists():
        data = json.loads(SYNC_LOG.read_text())
    else:
        data = {}
    # Defensive defaults so older logs from prior versions keep working
    data.setdefault("processed_message_ids", [])
    data.setdefault("attachment_hashes", {})    # sha256 → report_name
    data.setdefault("last_sync", None)
    return data


def save_sync_log(log_data: dict):
    SYNC_LOG.write_text(json.dumps(log_data, indent=2))


def _token_has_all_scopes(creds) -> bool:
    """True if the loaded token covers every scope in SCOPES.

    `Credentials.valid` only checks expiry — it doesn't notice when SCOPES
    expanded after the token was minted. We check explicitly so adding a new
    scope (e.g. drive.readonly on top of an older gmail-only token) forces
    a fresh OAuth consent screen instead of silently using the old token.
    """
    if not creds:
        return False
    have = set(creds.scopes or [])
    need = set(SCOPES)
    return need.issubset(have)


def gmail_authenticate(force_reauth: bool = False):
    """OAuth2 flow — opens browser on first run or when scopes expand.

    Three paths:
      1. force_reauth=True  → always run the browser flow (used by --auth)
      2. Valid token with all required scopes → reuse silently
      3. Expired token with refresh_token   → refresh in-place
      4. Missing token / insufficient scopes → browser flow + write new token.json
    """
    creds = None
    if TOKEN_FILE.exists() and not force_reauth:
        try:
            creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
        except Exception as e:
            log.warning(f"Could not load token.json ({e}); will re-auth.")
            creds = None

    needs_browser = force_reauth or not creds or not _token_has_all_scopes(creds)

    if (not needs_browser) and creds and not creds.valid:
        # Token has the right scopes but is expired — try a silent refresh.
        if creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                log.warning(f"Refresh-token call failed ({e}); falling back to browser flow.")
                needs_browser = True
        else:
            needs_browser = True

    if needs_browser:
        if not CREDS_FILE.exists():
            raise FileNotFoundError(
                f"credentials.json not found at {CREDS_FILE}\n"
                "Download it from Google Cloud Console → APIs & Services → Credentials "
                "(OAuth client ID → Desktop app)."
            )
        # Visible heads-up so the user knows what's about to happen — the
        # browser-open call is otherwise silent.
        print()
        print("─" * 70)
        print(" Browser-based Google sign-in required.")
        print(" A browser tab will open in a moment.")
        print(" Approve BOTH requested permissions:")
        for s in SCOPES:
            print(f"   • {s}")
        print(" Then come back here — the flow will continue automatically.")
        print("─" * 70)
        print()
        log.info("Launching OAuth browser flow…")
        flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_FILE), SCOPES)
        creds = flow.run_local_server(port=0, open_browser=True,
                                       prompt='consent')   # force consent screen so all scopes are shown
        log.info("OAuth flow complete ✓")

    TOKEN_FILE.write_text(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def do_auth():
    """`--auth` entry point. Always runs the OAuth browser flow and confirms
    afterwards. Used for first-time setup and for re-consenting after SCOPES
    expand (e.g. when Drive was added on top of Gmail-only).
    """
    if not GMAIL_AVAILABLE:
        print("✗ Google API packages not installed.")
        print("  pip3 install google-auth-oauthlib google-api-python-client")
        sys.exit(1)
    if not CREDS_FILE.exists():
        print(f"✗ credentials.json not found at {CREDS_FILE}")
        print("  Download OAuth client (Desktop app) from Google Cloud Console")
        print("  → APIs & Services → Credentials, then save it to that path.")
        sys.exit(2)

    print("Forcing fresh OAuth consent — current token.json (if any) will be replaced.")
    try:
        service = gmail_authenticate(force_reauth=True)
    except Exception as e:
        print(f"✗ Auth failed: {e}")
        sys.exit(3)

    # Sanity-check: query the user's own profile so we confirm the token works.
    try:
        prof = service.users().getProfile(userId="me").execute()
        print()
        print("✓ Authenticated as:", prof.get("emailAddress"))
        print(f"✓ Token saved to:   {TOKEN_FILE}")
        # Confirm the scopes that ended up on the token.
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
        print("✓ Scopes granted:")
        for s in creds.scopes or []:
            print(f"   • {s}")
        missing = set(SCOPES) - set(creds.scopes or [])
        if missing:
            print(f"⚠  Missing scopes: {missing}")
            print("   Re-run --auth and ensure you click 'Allow' on every permission prompt.")
            sys.exit(4)
    except Exception as e:
        print(f"⚠  Auth succeeded but profile call failed: {e}")


def download_attachment(service, message_id: str, attachment_id: str, dest_path: Path):
    """Download a Gmail attachment and save to dest_path. Returns (bytes_written, sha256)."""
    att = service.users().messages().attachments().get(
        userId="me", messageId=message_id, id=attachment_id
    ).execute()
    data = base64.urlsafe_b64decode(att["data"])
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    dest_path.write_bytes(data)
    sha = hashlib.sha256(data).hexdigest()
    log.info(f"  ✓ Saved {dest_path.name}  ({len(data):,} bytes, sha256={sha[:10]}…)")
    return len(data), sha


def _save_inline_attachment(part: dict, dest_path: Path):
    """For very small attachments Gmail inlines the base64 bytes in body.data
    instead of providing an attachmentId. Save those too."""
    data_b64 = (part.get("body") or {}).get("data")
    if not data_b64:
        return None
    data = base64.urlsafe_b64decode(data_b64)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    dest_path.write_bytes(data)
    sha = hashlib.sha256(data).hexdigest()
    log.info(f"  ✓ Saved (inline) {dest_path.name}  ({len(data):,} bytes, sha256={sha[:10]}…)")
    return len(data), sha


# MIME types that obviously AREN'T a CSV attachment, even if the filename happens
# to end in .csv (very rare, but defensive). Anything else is accepted as long
# as the filename ends in .csv — we trust the extension over Gmail's MIME tag,
# because senders mislabel CSVs with all kinds of MIME types in the wild.
NON_ATTACHMENT_MIMES = {
    "text/html",
    "text/plain",        # only excluded when there's no filename — see _looks_like_csv_part
    "multipart/alternative",
    "multipart/mixed",
    "multipart/related",
    "application/pgp-signature",
}


def _looks_like_csv_part(part: dict) -> bool:
    """Decide if a MIME part is a CSV attachment worth downloading.

    Strategy: trust the filename. If it ends in .csv (case-insensitive) and the
    part has a body, it's an attachment. Otherwise fall through.
    """
    fname = (part.get("filename") or "").strip()
    if not fname:
        return False
    if not fname.lower().endswith(".csv"):
        return False
    body = part.get("body") or {}
    return bool(body.get("attachmentId") or body.get("data"))


def _looks_like_xls_part(part: dict) -> bool:
    """Decide if a MIME part is an XLS/XLSX attachment worth downloading.

    Billing reports from Zoho arrive as .xls or .xlsx files.
    """
    fname = (part.get("filename") or "").strip().lower()
    if not fname:
        return False
    if not (fname.endswith(".xls") or fname.endswith(".xlsx")):
        return False
    body = part.get("body") or {}
    return bool(body.get("attachmentId") or body.get("data"))


def _get_header(message: dict, name: str) -> str:
    """Pull a header value (case-insensitive) from a Gmail message."""
    for h in message.get("payload", {}).get("headers", []):
        if h.get("name", "").lower() == name.lower():
            return h.get("value", "")
    return ""


def _expected_subject(report_name: str) -> str:
    return f'{SUBJECT_PREFIX}{report_name}{SUBJECT_SUFFIX}'


def _norm_subject(s: str) -> str:
    """Normalise an email subject so substring matching is robust.

    Real-world subjects vary in ways that break naive substring matching:
      - Smart quotes (" " ' ') instead of straight quotes
      - Reply/forward prefixes ('Fwd:', 'FW:', 'Re:', 'RE:', '[EXT]')
      - Trailing whitespace, zero-width characters, NBSP
      - Case differences in non-content words
    """
    if not s:
        return ""
    out = s
    # Strip common prefixes (handles chains like 'Fwd: Re: Fwd: …')
    while True:
        m = re.match(r"^\s*(?:re|fwd?|fw|\[ext\]|\[external\])\s*:\s*", out, re.IGNORECASE)
        if not m:
            break
        out = out[m.end():]
    # Replace smart quotes with straight ones
    SMART_TO_STRAIGHT = {
        "“": '"', "”": '"',     # double curly quotes
        "‘": "'", "’": "'",     # single curly quotes
        "«": '"', "»": '"',     # « »
        "′": "'", "″": '"',     # primes
    }
    for src, dst in SMART_TO_STRAIGHT.items():
        out = out.replace(src, dst)
    # Collapse whitespace; remove zero-width chars and NBSPs
    out = out.replace(" ", " ").replace("​", "")
    out = " ".join(out.split())
    return out.strip().lower()


def _subject_matches(subject: str, report_name: str) -> bool:
    """Is this email subject for the given report?

    We try, in order of strictness:
      1. The full 'Check out the "<name>" report' phrase (normalised)
      2. Just the inner report name (normalised) — handles senders who
         change the surrounding phrasing
    """
    nsubj = _norm_subject(subject)
    nfull = _norm_subject(_expected_subject(report_name))
    if nfull and nfull in nsubj:
        return True
    # Fallback — the distinctive inner phrase is itself quite specific
    return _norm_subject(report_name) in nsubj


def sync_gmail(force: bool = False) -> dict:
    """Search inbox for the 3 report emails, download new attachments only.

    Args:
        force: When True, ignore processed_message_ids — always grab the
            single newest message for each report. Use this for user-initiated
            'Sync Gmail' button clicks where the explicit intent is "give me
            the freshest data." The daily scheduled cron passes force=False so
            it doesn't re-download identical content on every run.

    De-dupe strategy is two-layered:
      1. processed_message_ids   → skip messages we've already opened (force=False only)
      2. attachment_hashes       → even on a new message, mark 'unchanged' if
                                   the file bytes match something we've already saved
                                   (informational; file is always written either way)

    Returns a status dict:
        {
            "ok":          bool,        # False = blocking failure (auth, packages, etc.)
            "error":       str | None,  # human-readable reason when ok=False
            "downloaded":  [report_name, …],  # reports we just refreshed from Gmail
            "unchanged":   [report_name, …],  # found a message but bytes were identical
            "missing":     [report_name, …],  # no matching email found at all
            "new_files":   bool,        # True if at least one CSV was actually replaced
        }
    """
    status = {"ok": True, "error": None, "downloaded": [], "unchanged": [],
              "missing": [], "new_files": False, "force": force,
              # report_name → ISO timestamp of the source email's internalDate
              "source_email_dates": {}}

    if not GMAIL_AVAILABLE:
        status.update(ok=False, error="Google API packages not installed (pip3 install "
                                       "google-auth-oauthlib google-api-python-client)")
        log.error(status["error"])
        return status

    if not CREDS_FILE.exists():
        status.update(ok=False, error=(
            f"credentials.json not found at {CREDS_FILE}. Download an OAuth client "
            "(Desktop app) from Google Cloud Console → APIs & Services → Credentials, "
            "save it to that path, then re-run."))
        log.error(status["error"])
        return status

    sync_data = load_sync_log()
    processed = set(sync_data["processed_message_ids"])
    seen_hashes = dict(sync_data["attachment_hashes"])     # sha256 → report_name

    try:
        service = gmail_authenticate()
        log.info("Gmail authenticated ✓")
    except Exception as e:
        status.update(ok=False, error=f"Gmail auth failed: {e}")
        log.error(status["error"])
        return status

    for report_name, dest_file in EMAIL_MAP.items():
        expected_subject = _expected_subject(report_name)
        is_xls_report    = report_name in XLS_EMAIL_MAP

        # Gmail's `subject:"…"` operator does substring matching, but the
        # exact subject contains literal double-quotes around the report name:
        #     Check out the "Incoming Order Count Insurer Wise" report
        # If we put that whole thing inside subject:"…" we end up with nested
        # quotes that Gmail's parser breaks on, so it returns AND-of-keywords
        # matches instead of the exact phrase. The cure: search for the
        # distinctive inner phrase only — then verify the full subject below.
        # CSV reports use filename:csv; billing XLS reports use filename:xls.
        file_filter = "filename:xls" if is_xls_report else "filename:csv"
        query = f'subject:"{report_name}" has:attachment {file_filter}'
        log.info(f"Searching: {query}")

        result = service.users().messages().list(
            userId="me", q=query, maxResults=10
        ).execute()
        msg_refs = result.get("messages", [])

        if not msg_refs:
            log.warning(f"  No emails found for: {report_name}")
            status["missing"].append(report_name)
            continue

        # Walk newest → oldest (Gmail returns newest first). Stop after we save one.
        # CRITICAL: only add msg_id to processed (the persistent blacklist) AFTER
        # a successful download. If we permanently blacklisted on every miss,
        # a single false-negative (e.g. transient subject-normalisation bug)
        # would silently make us skip that email forever.
        #
        # When force=True (user clicked "Sync Gmail" expecting fresh data),
        # ignore processed_message_ids entirely — always grab the newest
        # matching message. The on-disk file is overwritten with whatever the
        # newest email contains, regardless of whether we've seen that
        # message ID before.
        rejected_in_session = []    # for verbose logging only
        for ref in msg_refs:
            msg_id = ref["id"]
            if msg_id in processed and not force:
                log.debug(f"  Already processed msg {msg_id[:12]}… ({report_name})")
                continue

            message = service.users().messages().get(
                userId="me", id=msg_id, format="full"
            ).execute()
            subject = _get_header(message, "Subject").strip()
            from_h  = _get_header(message, "From").strip()

            if not _subject_matches(subject, report_name):
                # NOTE: do NOT add to `processed`. The subject check might be
                # wrong — if so, we want this message to be re-evaluated next run.
                rejected_in_session.append((msg_id, subject, "subject-mismatch"))
                log.info(f"  ✗ Subject mismatch [{msg_id[:12]}…] from={from_h!r}  subject={subject!r}")
                continue

            # Walk the MIME tree and find a CSV part. Trust the filename — Gmail's
            # MIME-type tag is unreliable in the wild.
            _part_fn   = _looks_like_xls_part if is_xls_report else _looks_like_csv_part
            _part_type = "XLS" if is_xls_report else "CSV"
            csv_parts  = [p for p in _iter_parts(message.get("payload", {})) if _part_fn(p)]
            if not csv_parts:
                # List every part so the user can see WHY no file was found
                seen = [(p.get("filename") or "(no filename)", p.get("mimeType") or "")
                        for p in _iter_parts(message.get("payload", {}))]
                log.info(f"  ✗ No {_part_type} attachment on msg [{msg_id[:12]}…]  parts={seen}")
                rejected_in_session.append((msg_id, subject, f"no-{_part_type.lower()}-attachment"))
                # Don't blacklist — sender might re-send with a real attachment.
                continue

            # Pick the LAST matching part. Multi-attachment emails are rare; when
            # they exist the more recent/relevant one tends to be later in the body.
            part = csv_parts[-1]
            fname = part.get("filename", "")

            log.info(f"  → Downloading: {fname!r} (msg {msg_id[:12]}…  subject={subject!r})")
            try:
                body = part.get("body") or {}
                att_id = body.get("attachmentId")
                if att_id:
                    result = download_attachment(service, msg_id, att_id, dest_file)
                else:
                    result = _save_inline_attachment(part, dest_file)
                if not result:
                    raise RuntimeError("attachment had no body data")
                _bytes, sha = result
            except Exception as e:
                log.error(f"  Download failed: {e}")
                rejected_in_session.append((msg_id, subject, f"download-error: {e}"))
                continue

            # Record the source email's date — Gmail's internalDate is ms since epoch.
            try:
                internal_ms = int(message.get("internalDate", 0))
                if internal_ms:
                    iso = datetime.fromtimestamp(internal_ms/1000, tz=timezone.utc).isoformat()
                    status["source_email_dates"][report_name] = iso
            except Exception:
                pass

            # Hash-level de-dupe: if the same bytes were saved for the same
            # report previously, the input is unchanged — don't flag as new.
            if seen_hashes.get(sha) == report_name:
                log.info(f"  = Same content as prior sync — input unchanged")
                status["unchanged"].append(report_name)
            else:
                seen_hashes[sha] = report_name
                status["downloaded"].append(report_name)
                status["new_files"] = True
                log.info(f"  ✓ New content for {report_name}")

            processed.add(msg_id)
            break    # got the latest copy for this report — move to next report

        # If we walked all msg_refs and none yielded a saved attachment, log a
        # detailed reason so the user can see which emails were rejected and why.
        if (report_name not in status["downloaded"] and
            report_name not in status["unchanged"]):
            status["missing"].append(report_name)
            if rejected_in_session:
                log.warning(
                    f"  {report_name}: {len(rejected_in_session)} message(s) seen but none usable:"
                )
                for mid, subj, reason in rejected_in_session[:5]:
                    log.warning(f"    [{mid[:12]}…] {reason}  subject={subj!r}")

    # ── Extra pass: files that share a subject with another report ───────────
    # Daily_Insurer_Billing_2.xls arrives in the SAME "Daily Insurer Billing"
    # email thread as Daily_Insurer_Billing.xls.  The main loop above can only
    # pick one destination per subject, so we scan those same messages a second
    # time and route by attachment filename.
    if BILLING_EXTRA_FILES:
        q_extra = 'subject:"Daily Insurer Billing" has:attachment filename:xls'
        try:
            extra_refs = service.users().messages().list(
                userId="me", q=q_extra, maxResults=20
            ).execute().get("messages", [])
        except Exception as exc:
            log.warning(f"  BILLING_EXTRA_FILES search failed: {exc}")
            extra_refs = []

        for extra_fname, extra_dest in BILLING_EXTRA_FILES.items():
            found = False
            for ref in extra_refs:
                try:
                    msg_e = service.users().messages().get(
                        userId="me", id=ref["id"], format="full"
                    ).execute()
                except Exception:
                    continue
                for part in _iter_parts(msg_e.get("payload", {})):
                    part_name = (part.get("filename") or "").strip().lower()
                    if part_name == extra_fname:
                        body = part.get("body") or {}
                        att_id = body.get("attachmentId")
                        if att_id:
                            try:
                                res = download_attachment(
                                    service, ref["id"], att_id, extra_dest
                                )
                                if res:
                                    _bytes, sha = res
                                    label = extra_dest.name
                                    if seen_hashes.get(sha) == label:
                                        log.info(f"  = {label}: content unchanged")
                                        status["unchanged"].append(label)
                                    else:
                                        seen_hashes[sha] = label
                                        status["downloaded"].append(label)
                                        status["new_files"] = True
                                        log.info(f"  ✓ Downloaded {label} (billing multi-file)")
                                    found = True
                            except Exception as exc:
                                log.warning(f"  {extra_dest.name} download failed: {exc}")
                        break
                if found:
                    break
            if not found:
                log.warning(f"  {extra_dest.name}: not found in billing thread")
                status["missing"].append(extra_dest.name)

    sync_data["processed_message_ids"] = list(processed)
    sync_data["attachment_hashes"]     = seen_hashes
    sync_data["last_sync"]             = datetime.now(timezone.utc).isoformat()
    save_sync_log(sync_data)
    return status


def _iter_parts(payload):
    """Recursively yield MIME parts so we catch attachments nested under multipart/mixed."""
    if not payload:
        return
    yield payload
    for sub in payload.get("parts", []) or []:
        yield from _iter_parts(sub)


# ══════════════════════════════════════════════════════════════════════════════
# 2.  DATA PROCESSING  (reads all Excel files → dict)
# ══════════════════════════════════════════════════════════════════════════════

_CURRENCY_PREFIX_RE = re.compile(
    r"^\s*(?:INR|Rs\.?|₹|\$|USD|EUR|€|£|GBP)\s*",
    re.IGNORECASE,
)


def _clean_number(v):
    """Convert a possibly-formatted number string to a plain numeric string.

    Real-world CSV cells the script must handle:
        '627,894.00'                  → '627894.00'
        '1,23,456'   (Indian format)  → '123456'
        'INR 627,894.00'              → '627894.00'
        '₹ 1,234.56'                  → '1234.56'
        'Rs. 1,234'                   → '1234'
        '(1,234)'    (accounting neg) → '-1234'
        '  1234.5  '                  → '1234.5'
        ''                            → ''
        None                          → ''
    Returns '' if the input has no numeric content.
    """
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return ""
    # Strip currency prefixes
    s = _CURRENCY_PREFIX_RE.sub("", s)
    # Accounting-style negatives: '(1,234)' → '-1,234'
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    # Remove thousand separators (handles both 1,234,567 and 1,23,456)
    s = s.replace(",", "").replace(" ", "")
    return s.strip()


def safe_float(v, default=0.0):
    s = _clean_number(v)
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def safe_int(v, default=0):
    s = _clean_number(v)
    if not s:
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


_MONTH_ORDER = {
    "jan":1, "feb":2, "mar":3, "apr":4, "may":5, "jun":6,
    "jul":7, "aug":8, "sep":9, "oct":10,"nov":11,"dec":12,
}

def _month_sort_key(label: str) -> tuple:
    """Sort 'Mar 2025', 'Apr 2025', … chronologically. Unknowns sort last."""
    s = (label or "").strip().lower()
    parts = s.split()
    if len(parts) >= 2:
        m = _MONTH_ORDER.get(parts[0][:3], 99)
        try:
            y = int(parts[1])
        except ValueError:
            y = 9999
        return (y, m)
    return (9999, 99)


def _normalize_month(raw: str, fallback_name: str = "") -> str:
    """Normalise a month value to 'May', 'June', or 'Q2'.

    Accepts:
      - 'May', 'May 2026', 'May-26', 'may'              → 'May'
      - 'June', 'Jun 2026', 'June 26', 'jun'            → 'June'
      - datetime stringified ('2026-05-01 00:00:00')    → 'May'
      - empty / unknown → parse 'May' or 'June' out of fallback_name; else 'Q2'
    """
    s = (raw or "").strip().lower()
    if s:
        if "may" in s or s.startswith("05") or "-05-" in s or "/05/" in s:
            return "May"
        if "jun" in s or s.startswith("06") or "-06-" in s or "/06/" in s:
            return "June"
        if "apr" in s or s.startswith("04") or "-04-" in s or "/04/" in s:
            return "April"
    # Fallback: pull from the row name (existing convention: 'X - May', 'Y - June')
    n = (fallback_name or "").lower()
    if " - may" in n or n.endswith("may"):  return "May"
    if " - june" in n or n.endswith("june") or " - jun" in n: return "June"
    # No month info anywhere → quarter-wide
    return "Q2"


def _parse_monthly_targets(rows: list) -> dict:
    """Walk the Plan sheet for '<Month> 2026 Month' section blocks.

    Each block has a header row like:
        ('May 2026 Month', None, None, ...)
    followed by a column-header row with 'Total Incoming orders to achieve'
    and 'Total closed orders to achieve', then per-insurer rows, then a
    trailing total row with mostly Nones.

    Returns:
        {"May 2026": {"incoming_required": N, "closed_required": N},
         "Jun 2026": {...}}
    """
    targets = {}
    # Tag JSON keys consistently with the rest of the dataset (3-letter month).
    SECTION_MAP = {
        "may 2026 month":  "May 2026",
        "june 2026 month": "Jun 2026",
        "jun 2026 month":  "Jun 2026",
        "april 2026 month":"Apr 2026",
        "apr 2026 month":  "Apr 2026",
    }

    i = 0
    while i < len(rows):
        first = rows[i]
        head = str(first[0] or "").strip().lower() if first else ""
        if head in SECTION_MAP:
            json_key = SECTION_MAP[head]
            # Next row should be column headers
            if i + 1 >= len(rows):
                break
            hdr_row = rows[i + 1]
            hdr = [str(c or "").strip().lower() for c in hdr_row]

            def hcol(*needles):
                for j, h in enumerate(hdr):
                    if all(n in h for n in needles):
                        return j
                return None

            inc_col    = hcol("incoming", "achieve")
            closed_col = hcol("closed",   "achieve")
            if inc_col is None and closed_col is None:
                i += 1
                continue

            inc_total = 0.0
            cl_total  = 0.0
            per_insurer = {}
            j = i + 2
            while j < len(rows):
                row = rows[j]
                name = str(row[0] or "").strip() if row else ""
                # Stop on a blank line OR when we hit the next section header
                if not name:
                    j += 1
                    if not any(c for c in (row or [])):    # truly empty → end
                        break
                    continue
                if name.lower() in SECTION_MAP:
                    break
                if "total" in name.lower() or "target" in name.lower():
                    j += 1
                    continue
                inc_val = safe_float(row[inc_col]) if (inc_col is not None and inc_col < len(row)) else 0
                cl_val  = safe_float(row[closed_col]) if (closed_col is not None and closed_col < len(row)) else 0
                # Skip rows where both required values are 0 — they're noise
                if inc_val > 0 or cl_val > 0:
                    per_insurer[name] = {
                        "incoming_required": int(round(inc_val)),
                        "closed_required":   int(round(cl_val)),
                    }
                    inc_total += inc_val
                    cl_total  += cl_val
                j += 1

            targets[json_key] = {
                "incoming_required": int(round(inc_total)),
                "closed_required":   int(round(cl_total)),
                "by_insurer":        per_insurer,
            }
            i = j
            continue
        i += 1

    return targets


def read_quarter_target() -> dict:
    """Parse Quarter Target.xlsx → structured dict."""
    if not QUARTER_TARGET_FILE.exists():
        log.warning("Quarter Target.xlsx not found")
        return {}

    wb = openpyxl.load_workbook(str(QUARTER_TARGET_FILE), data_only=True)
    data = {}

    # ── April 2026 actuals ────────────────────────────────────────────────────
    # The "Apr 2026" sheet has MULTIPLE stacked tables — only the first one
    # (rows 2..N) is customer-by-customer revenue. Below that there's a Q2
    # target summary, then an "April Revenue (in Lakhs) vs Target" block
    # that uses the SAME insurer names but values in lakhs/percent. If we
    # don't stop at the first blank row, those rows leak into apr_customers
    # as 'Aditya Birla rev=₹18 conv=2700%' style junk.
    if "Apr 2026" in wb.sheetnames:
        ws = wb["Apr 2026"]
        rows = list(ws.values)
        apr_customers = []
        SECTION_BREAK_KEYWORDS = ("q2 fy", "q2 total", "month", "in lakhs",
                                  "april revenue", "revenue target", "initiative",
                                  "total incremental")

        for row in rows[1:]:
            # 1) Hard stop: a fully blank row is the section separator. We
            #    use this even before checking row[0], because row[0] can be
            #    None on a "totals" row that we want to skip but NOT stop on.
            if not any(c is not None and str(c).strip() not in ('','nan') for c in row):
                break

            # 2) Skip totals/blank-name rows (row 9 = totals: name is None but
            #    numbers are filled in)
            if not row[0] or str(row[0]).strip() in ("", "nan"):
                continue
            name = str(row[0]).strip()

            # 3) Section-header text → we've left the customer block, stop.
            lname = name.lower()
            if any(k in lname for k in SECTION_BREAK_KEYWORDS):
                break

            # 4) Missing or zero revenue → skip (not a section break, just an
            #    empty row mid-block).
            if not row[2]:
                continue
            rev = safe_float(row[2])
            if rev <= 0:
                continue

            # 5) Sanity guard: the upper customer block always has revenue
            #    in actual rupees (lakhs+). The lookalike block below uses
            #    values in lakhs (single/double digits). Anything under ₹1000
            #    is suspicious — either a different unit or a stray number —
            #    so we drop it. (Cardiotrack Home Services with rev=₹1298 is
            #    above the threshold, so it stays.)
            if rev < 1000:
                continue
            # Same defence for conversion — the real conversion column is a
            # 0..1 fraction. The lakhs-block has integers like 27 in the same
            # position. Drop anything > 1.5 (allow tiny rounding error).
            conv = safe_float(row[5])
            if conv > 1.5:
                continue

            apr_customers.append({
                "insurer":      name,
                "closed_cases": safe_int(row[1]),
                "revenue":      rev,
                "avg_order":    safe_float(row[3]),
                "incoming":     safe_int(row[4]),
                "conversion":   conv,
            })
        data["apr_customers"] = apr_customers
        data["apr_total_revenue"] = sum(c["revenue"] for c in apr_customers)

    # ── Plan / Pipeline ───────────────────────────────────────────────────────
    if "Plan" in wb.sheetnames:
        ws = wb["Plan"]
        rows = list(ws.values)
        header = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]

        # Resolve column positions by header name. This makes the parser robust
        # to columns being inserted, moved, or renamed (e.g. when 'Month' was
        # added between Type and Key Action).
        def col(*names, default=None):
            for n in names:
                if n in header:
                    return header.index(n)
            return default

        NAME_COL    = col("insurer name", default=0)
        DELTA_COL   = col("delta 50 lakh", "delta", default=1)
        PROB_COL    = col("probability", default=2)
        WEIGHT_COL  = col("revenue based on probability ₹", "weighted revenue", "revenue", default=3)
        CLOSED_COL  = col("additional closed orders", default=4)
        INC_COL     = col("additional incoming orders", default=5)
        TYPE_COL    = col("type", default=6)
        MONTH_COL   = col("month", "target month", "delta month")           # may be None
        ACTION_COL  = col("key action", "action", "next action", default=7) # shifts if Month is inserted

        def get(row, idx):
            return row[idx] if idx is not None and idx < len(row) else None

        pipeline = []
        for row in rows[1:]:
            name_v = get(row, NAME_COL)
            if not name_v or str(name_v).strip() in ("", "nan"):
                continue
            name = str(name_v).strip()
            if "Total" in name or "Target" in name or "Pipeline" in name:
                continue
            delta = safe_float(get(row, DELTA_COL))
            prob  = safe_float(get(row, PROB_COL))
            if delta <= 0:
                continue

            # ── Resolve the month (May / June / Q2) ─────────────────────────
            month_raw = ""
            m_val = get(row, MONTH_COL)
            if m_val:
                month_raw = str(m_val).strip()
            month = _normalize_month(month_raw, fallback_name=name)

            type_v   = get(row, TYPE_COL)
            action_v = get(row, ACTION_COL)
            pipeline.append({
                "name":        name,
                "month":       month,            # ← May | June | Q2
                "delta":       delta,
                "probability": prob,
                "weighted":    safe_float(get(row, WEIGHT_COL)),
                "add_closed":  safe_float(get(row, CLOSED_COL)),
                "add_incoming":safe_float(get(row, INC_COL)),
                "type":        str(type_v).strip() if type_v else "Unknown",
                "key_action":  str(action_v).strip() if action_v else "",
            })
        data["pipeline"] = pipeline
        data["may_pipeline_total"]  = sum(p["weighted"] for p in pipeline if p["month"] == "May")
        data["june_pipeline_total"] = sum(p["weighted"] for p in pipeline if p["month"] == "June")
        data["q2_pipeline_total"]   = sum(p["weighted"] for p in pipeline if p["month"] == "Q2")
        data["plan_has_month_col"]  = MONTH_COL is not None

        # ── May/June "required orders" targets ────────────────────────────
        # The Plan sheet has dedicated sections lower down ("May 2026 Month",
        # "June 2026 Month") with columns 'Total Incoming orders to achieve'
        # and 'Total closed orders to achieve' per insurer. Sum them to get
        # the headline target numbers shown on the dashboard.
        data["monthly_targets"] = _parse_monthly_targets(rows)

    # ── Action items (Sheet2 columns H-L) ────────────────────────────────────
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
        actions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Cols H,I,J,K,L → indices 7,8,9,10,11
            if len(row) < 9:
                continue
            initiative = row[7] if len(row) > 7 else None
            potential  = row[8] if len(row) > 8 else None
            deadline   = row[9] if len(row) > 9 else None
            owner      = row[10] if len(row) > 10 else None
            if not initiative or not str(initiative).strip():
                continue
            init_text = str(initiative).strip()
            # Skip the summary row at the bottom
            if "Total incremental" in init_text or "Total" == init_text:
                continue

            # Owner-or-action lives in col K; sometimes deadline text is here too
            owner_text = str(owner).strip() if owner else ""

            # Format the deadline if it's a datetime
            if hasattr(deadline, "strftime"):
                deadline_text = deadline.strftime("%d %b %Y")
            else:
                deadline_text = str(deadline).strip() if deadline else ""

            # Best-effort month tag for the badge — pulled from initiative text
            month_tag = "Q2"
            lower = init_text.lower()
            if "may" in lower:  month_tag = "May"
            elif "jun" in lower: month_tag = "June"
            elif "apr" in lower: month_tag = "Apr"
            elif hasattr(deadline, "month"):
                month_tag = {4:"Apr",5:"May",6:"June"}.get(deadline.month, "Q2")

            actions.append({
                "initiative": init_text,
                "potential":  str(potential).strip() if potential else "—",
                "deadline":   deadline_text or "—",
                "owner":      owner_text or "—",
                "month":      month_tag,
                "status":     "Open",      # no status column in source — default
            })
        data["key_actions"] = actions

    # ── Overall Billing (monthly trend) ──────────────────────────────────────
    if "Overall Billing Data" in wb.sheetnames:
        ws = wb["Overall Billing Data"]
        monthly = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                monthly[str(row[0]).strip()] = safe_float(row[1])
        data["monthly_billing"] = monthly

    # ── Insurer-wise billing ──────────────────────────────────────────────────
    if "BIlling done Insurer Wise" in wb.sheetnames:
        ws = wb["BIlling done Insurer Wise"]
        insurer_monthly = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:
                month   = str(row[1]).strip()
                insurer = str(row[2]).strip()
                rev     = safe_float(row[0])
                insurer_monthly.setdefault(insurer, {})[month] = rev
        data["insurer_monthly_billing"] = insurer_monthly

    return data


def _read_csv_rows(path: Path):
    """Yield header + data rows from a CSV file.

    Handles:
      - UTF-8 with or without BOM
      - Trailing whitespace in values
      - Blank trailing lines
      - openpyxl-saved CSVs (which use \\r\\n)
    """
    # 'utf-8-sig' strips a leading BOM if present; falls back fine otherwise.
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        # Sniff the dialect — Excel sometimes exports semicolon-separated.
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        for row in reader:
            # Strip whitespace on every cell so " May 2026" / "May 2026 " match
            yield [c.strip() if isinstance(c, str) else c for c in row]


def _ffill_month(records: list) -> list:
    """Forward-fill the 'month' column.

    Source CSVs (just like the Excel originals) only write the month on the
    first row of each group — subsequent rows in the same month leave it blank.
    A naive reader drops every blank-month row; this restores them.
    """
    last_month = ""
    out = []
    for r in records:
        m = (r.get("month") or "").strip()
        if m:
            last_month = m
        elif last_month:
            r["month"] = last_month
        out.append(r)
    return out


def read_orders_file(path: Path, label: str) -> list:
    """Read Incoming / Closed order count CSV → list of {month, insurer, count}.

    Expected columns (header row): Month, Insurer Name, Count
    Column order is detected by header text so reordering is safe.
    """
    if not path.exists():
        log.warning(f"{label} file not found: {path.name}")
        return []

    rows = list(_read_csv_rows(path))
    if not rows:
        log.warning(f"{label} CSV is empty: {path.name}")
        return []

    header = [h.lower() for h in rows[0]]
    # Find columns by name with sensible fallbacks for known header variants.
    def find(*keywords, default=None):
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return default

    MONTH_COL   = find("month", default=0)
    INSURER_COL = find("insurer", default=1)
    COUNT_COL   = find("count", default=2)

    records = []
    for row in rows[1:]:
        if not any(c for c in row):                # skip fully blank lines
            continue
        if len(row) <= max(MONTH_COL, INSURER_COL, COUNT_COL):
            continue
        insurer = row[INSURER_COL]
        if not insurer:                            # need at least an insurer
            continue
        records.append({
            "month":   row[MONTH_COL],
            "insurer": insurer,
            "count":   safe_int(row[COUNT_COL]),
        })
    records = _ffill_month(records)
    log.info(f"  {label}: {len(records)} rows read")
    return records


def read_revenue_file(path: Path) -> list:
    """Read Revenue_Generated_Insurer_Wise.csv → list of {month, insurer, amount}.

    Expected columns: Month, Insurer Name, Amount/Revenue
    """
    if not path.exists():
        log.warning(f"Revenue file not found: {path.name}")
        return []

    rows = list(_read_csv_rows(path))
    if not rows:
        log.warning(f"Revenue CSV is empty: {path.name}")
        return []

    header = [h.lower() for h in rows[0]]
    def find(*keywords, default=None):
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return default

    MONTH_COL   = find("month", default=0)
    INSURER_COL = find("insurer", default=1)
    # Revenue file's amount column is often labelled 'Total Total Amount to be
    # Billed' or 'Amount' or 'Revenue'. Match any of those.
    AMT_COL     = find("amount", "billed", "revenue", default=2)

    records = []
    for row in rows[1:]:
        if not any(c for c in row):
            continue
        if len(row) <= max(MONTH_COL, INSURER_COL, AMT_COL):
            continue
        insurer = row[INSURER_COL]
        if not insurer:
            continue
        records.append({
            "month":   row[MONTH_COL],
            "insurer": insurer,
            "amount":  safe_float(row[AMT_COL]),
        })
    records = _ffill_month(records)
    log.info(f"  Revenue: {len(records)} rows read")
    return records


# ── Canonical insurer name (mirrors JS _canonicalName in dashboard) ──────────
_MONTH_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
_MONTH_NUM  = {v.lower():k for k,v in _MONTH_ABBR.items()}

def _canonical_insurer_name(n: str) -> str:
    if n == 'Acko Life':                                          return 'Acko Life Insurance'
    if n in ('Acko Health - Insurance Limited',
             'Acko Health -  Insurance Limited', 'Acko General'): return 'Acko Health Insurance'
    if n in ('Bajaj Life Insurance Company',
             'Bajaj Allianz Life Insurance Company'):             return 'Bajaj Life'
    if n == 'Tata AIA Life':                                      return 'TATA AIA Life Insurance Company'
    if n == 'SBI Life':                                           return 'SBI LIfe Insurance Company'
    if n in ('TAIG Health Insurance Company Ltd.',
             'TATA AIG Health Insurance Company Ltd.',
             'TATA AIG General Insurance'):                        return 'TATA AIG Health Insurance'
    if n == 'A8228Mph f7Rq7v0X':                                  return 'HDFC Life'
    if n == 'M7P4hW7Z Z9xz93cY':                                  return 'Star Union Daichi Life Insurance'
    if n == 'nA5N724T IpSA357U':                                   return 'TATA AIA Life Insurance Company'
    if n == '5MyHaB07 3VrA426d':                                   return 'SBI LIfe Insurance Company'
    return n


def _parse_date_to_month_label(s: str) -> str:
    """Convert various date formats to 'Mmm YYYY', e.g. '2026-04-15' → 'Apr 2026'."""
    s = (s or "").strip()
    # ISO: 2026-04-15
    m = re.match(r"(\d{4})-(\d{1,2})-\d{1,2}", s)
    if m:
        return f"{_MONTH_ABBR[int(m.group(2))]} {m.group(1)}"
    # DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        mn = int(m.group(2))
        if 1 <= mn <= 12:
            return f"{_MONTH_ABBR[mn]} {m.group(3)}"
    # Already "Apr 2026" / "April 2026"
    m = re.match(r"([A-Za-z]{3,9})\s+(\d{4})", s)
    if m:
        key = m.group(1)[:3].lower()
        if key in _MONTH_NUM:
            return f"{_MONTH_ABBR[_MONTH_NUM[key]]} {m.group(2)}"
    return s  # pass through unchanged


def read_billing_file(path: Path) -> list:
    """Read Daily_Insurer_Billing_Data.xls/.xlsx → list of {month_label, insurer, amount}.

    The file is distributed with a .xls extension but is actually an Office Open
    XML workbook (xlsx/ZIP format).  openpyxl handles it fine when loaded via
    BytesIO — bypassing the extension check that would otherwise reject it.

    Columns: Insurer Name | Policy Number | Order Closure Date | Total Amount to be Billed
    Date values are already datetime objects from openpyxl; we convert to 'Mmm YYYY'.
    """
    import io as _io

    if not path.exists():
        log.warning(f"Billing file not found: {path.name}")
        return []

    if not EXCEL_AVAILABLE:
        log.warning("openpyxl not available — cannot read billing file")
        return []

    try:
        with open(path, "rb") as fh:
            raw = _io.BytesIO(fh.read())
        wb = openpyxl.load_workbook(raw, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        log.warning(f"Could not open billing workbook {path.name}: {e}")
        return []

    if not all_rows:
        log.warning(f"Billing file is empty: {path.name}")
        return []

    # ── Column detection (keyword-based, case-insensitive) ────────────────────
    header = [str(h).lower() if h is not None else "" for h in all_rows[0]]

    def find(*keywords, default=None):
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return default

    DATE_COL    = find("date", "day", default=2)
    INSURER_COL = find("insurer", "company", "client", "partner", default=0)
    AMT_COL     = find("amount", "billing", "revenue", "billed", "total", default=3)

    max_col = max(DATE_COL, INSURER_COL, AMT_COL)
    records = []
    for row in all_rows[1:]:
        if not row or len(row) <= max_col:
            continue
        if not any(row):
            continue

        insurer_raw = row[INSURER_COL]
        date_val    = row[DATE_COL]
        amt_val     = row[AMT_COL]

        if insurer_raw is None or amt_val is None:
            continue

        insurer = _canonical_insurer_name(str(insurer_raw).strip())
        if not insurer:
            continue

        # openpyxl returns datetime objects for date cells
        if isinstance(date_val, datetime):
            month_label = date_val.strftime("%b %Y")
        else:
            month_label = _parse_date_to_month_label(str(date_val))

        amount = safe_float(str(amt_val)) if not isinstance(amt_val, (int, float)) else float(amt_val)
        records.append({"month_label": month_label, "insurer": insurer, "amount": amount})

    log.info(f"  Billing: {len(records)} daily rows → "
             f"{len(set(r['month_label'] for r in records))} months")
    return records


def _compute_yoy(billing_records: list) -> dict:
    """Compute Q2 YTD current-year vs prior-year comparison from billing records.

    Q2 = April + May + June.  Only compares months present in the current year
    (so Apr+May 2026 is matched against Apr+May 2025 — not all of Q2 2025).
    """
    from collections import defaultdict

    # Aggregate by (year, month_num, canonical_insurer)
    data: dict = defaultdict(float)
    for r in billing_records:
        lbl = r.get("month_label", "")
        parts = lbl.split()
        if len(parts) != 2:
            continue
        mon_key = parts[0][:3].lower()
        mon_num = _MONTH_NUM.get(mon_key, 0)
        if mon_num not in (4, 5, 6):          # Q2 only
            continue
        try:
            year = int(parts[1])
        except ValueError:
            continue
        data[(year, mon_num, r["insurer"])] += r["amount"]

    if not data:
        return {}

    current_year = max(k[0] for k in data)
    prior_year   = current_year - 1

    # Months that exist for the current year
    cy_months = sorted({k[1] for k in data if k[0] == current_year})

    cy_by_ins: dict = defaultdict(float)
    py_by_ins: dict = defaultdict(float)
    for (yr, mn, ins), amt in data.items():
        if mn not in cy_months:
            continue
        if yr == current_year:
            cy_by_ins[ins] += amt
        elif yr == prior_year:
            py_by_ins[ins] += amt

    all_ins = sorted(set(list(cy_by_ins) + list(py_by_ins)))
    cy_total = sum(cy_by_ins.values())
    py_total = sum(py_by_ins.values())
    abs_total = cy_total - py_total
    pct_total = round(abs_total / py_total * 100, 1) if py_total else None

    rows = []
    for ins in all_ins:
        cy  = cy_by_ins.get(ins, 0)
        py  = py_by_ins.get(ins, 0)
        abg = cy - py
        gp  = round(abg / py * 100, 1) if py else None
        rows.append({
            "insurer":          ins,
            "cy_revenue":       round(cy,  2),
            "py_revenue":       round(py,  2),
            "abs_growth":       round(abg, 2),
            "growth_pct":       gp,
            "contribution_pct": round(cy / cy_total * 100, 1) if cy_total else 0,
        })
    rows.sort(key=lambda r: r["cy_revenue"], reverse=True)

    gainers   = sorted([r for r in rows if r["abs_growth"] > 0],
                        key=lambda r: r["abs_growth"], reverse=True)[:3]
    decliners = sorted([r for r in rows if r["abs_growth"] < 0],
                        key=lambda r: r["abs_growth"])[:3]

    cy_months_str = [f"{_MONTH_ABBR[m]} {current_year}" for m in cy_months]

    return {
        "current_year":      current_year,
        "prior_year":        prior_year,
        "cy_months":         cy_months_str,
        "cy_total":          round(cy_total,  2),
        "py_total":          round(py_total,  2),
        "overall_growth":    round(abs_total, 2),
        "overall_growth_pct": pct_total,
        "by_insurer":        rows,
        "top_gainers":       gainers,
        "top_decliners":     decliners,
    }


def _revenue_kpis_from_detailed(records: list) -> dict:
    """Derive all revenue aggregations from the detailed billing records.

    Daily_Insurer_Billing.xls + Daily_Insurer_Billing_2.xls are the single
    source of truth for every revenue metric in the dashboard.

    Returns a dict with:
        apr_rev, may_rev, jun_rev, q2_total   — scalar grand totals
        monthly_billing                         — {month_str: grand_total}
        q2_by_insurer                           — {insurer: Q2_total}
        apr_by_insurer                          — {insurer: Apr_2026_total}
        may_by_insurer                          — {insurer: May_2026_total}
        jun_by_insurer                          — {insurer: Jun_2026_total}
        all_time_by_insurer                     — {insurer: all_time_total}
    """
    from collections import defaultdict as _dd

    monthly:  dict = _dd(float)
    apr_ins:  dict = _dd(float)
    may_ins:  dict = _dd(float)
    jun_ins:  dict = _dd(float)
    q2_ins:   dict = _dd(float)
    all_ins:  dict = _dd(float)

    for r in records:
        m   = r["month"]    # "Apr 2026", "Mar 2025", …
        ins = r["insurer"]
        tot = r["total"]
        monthly[m]  += tot
        all_ins[ins] += tot
        if m == "Apr 2026":
            apr_ins[ins] += tot
            q2_ins[ins]  += tot
        elif m == "May 2026":
            may_ins[ins] += tot
            q2_ins[ins]  += tot
        elif m == "Jun 2026":
            jun_ins[ins] += tot
            q2_ins[ins]  += tot

    apr_rev  = sum(apr_ins.values())
    may_rev  = sum(may_ins.values())
    jun_rev  = sum(jun_ins.values())
    q2_total = apr_rev + may_rev + jun_rev

    # Per-insurer revenue by month (for trend chart)
    monthly_by_ins: dict = _dd(lambda: _dd(float))
    for r in records:
        monthly_by_ins[r["month"]][r["insurer"]] += r["total"]
    # Convert nested defaultdicts to plain dicts for JSON serialisation
    monthly_by_ins_plain = {
        m: {k: round(v, 2) for k, v in ins_map.items()}
        for m, ins_map in monthly_by_ins.items()
    }

    return {
        "apr_rev":              round(apr_rev,  2),
        "may_rev":              round(may_rev,  2),
        "jun_rev":              round(jun_rev,  2),
        "q2_total":             round(q2_total, 2),
        "monthly_billing":      {m: round(v, 2) for m, v in monthly.items()},
        "monthly_by_insurer":   monthly_by_ins_plain,
        "q2_by_insurer":        {k: round(v, 2) for k, v in q2_ins.items()},
        "apr_by_insurer":       {k: round(v, 2) for k, v in apr_ins.items()},
        "may_by_insurer":       {k: round(v, 2) for k, v in may_ins.items()},
        "jun_by_insurer":       {k: round(v, 2) for k, v in jun_ins.items()},
        "all_time_by_insurer":  {k: round(v, 2) for k, v in all_ins.items()},
    }


def build_dashboard_data() -> dict:
    """Combine all sources into one JSON-serialisable dict."""
    log.info("Building dashboard data…")
    qt  = read_quarter_target()
    inc = read_orders_file(INCOMING_FILE, "Incoming Orders")
    cl  = read_orders_file(CLOSED_FILE,   "Closed Orders")

    # ── Single revenue source: Daily_Insurer_Billing*.xls ────────────────────
    # Detailed billing files are loaded first.  The summary billing file
    # (Daily_Insurer_Billing_Data.xls) is only read when the detailed files are
    # absent — avoids a redundant 12-second file read in the normal path.
    detailed_records = read_detailed_billing_files()
    if detailed_records:
        bk = _revenue_kpis_from_detailed(detailed_records)
        apr_rev         = bk["apr_rev"]
        may_rev         = bk["may_rev"]
        jun_rev         = bk["jun_rev"]
        q2_total        = bk["q2_total"]
        monthly_billing = bk["monthly_billing"]
        rev_by_insurer  = bk["q2_by_insurer"]      # Q2 per-insurer (replaces CSV)
        apr_rev_by_ins  = bk["apr_by_insurer"]      # Apr per-insurer
        lt_by_ins       = bk["all_time_by_insurer"] # all-time per-insurer
        billing_analysis_data = build_billing_analysis(detailed_records)
        # YoY source: convert to the format _compute_yoy expects
        yoy_src = [
            {"month_label": r["month"], "insurer": r["insurer"], "amount": r["total"]}
            for r in detailed_records
        ]
    else:
        # ── Fallback: summary billing file + Revenue CSV ──────────────────────
        log.warning("Detailed billing files absent — falling back to summary billing / Revenue CSV")
        billing = read_billing_file(BILLING_FILE)   # read only when needed
        rev     = read_revenue_file(REVENUE_FILE)
        bk      = None
        rev_q2  = [r for r in rev if r.get("month") in {"Apr 2026","May 2026","Jun 2026"}]
        csv_mb  = {}
        for r in rev:
            m = (r.get("month") or "").strip()
            if m: csv_mb[m] = csv_mb.get(m, 0) + r["amount"]
        qt_mb = qt.get("monthly_billing", {})
        def month_rev(month, fallback=0.0):
            if month in csv_mb and csv_mb[month] > 0: return csv_mb[month]
            if month in qt_mb  and qt_mb[month]  > 0: return qt_mb[month]
            return fallback
        apr_rev  = month_rev("Apr 2026", 4952222)
        may_rev  = month_rev("May 2026", 0)
        jun_rev  = month_rev("Jun 2026", 0)
        q2_total = apr_rev + may_rev + jun_rev
        monthly_billing = dict(qt_mb)
        for m, v in csv_mb.items():
            if v > 0: monthly_billing[m] = v
        def _grp(recs, key):
            out = {}
            for r in recs:
                ins = _canonical_insurer_name(r["insurer"])
                out[ins] = out.get(ins, 0) + r[key]
            return out
        rev_by_insurer  = _grp(rev_q2, "amount")
        apr_rev_by_ins  = _grp([r for r in rev if r.get("month") == "Apr 2026"], "amount")
        lt_by_ins       = _billing_lifetime(billing) if billing else {}
        billing_analysis_data = {}
        yoy_src = billing or [
            {"month_label": r["month"], "insurer": _canonical_insurer_name(r["insurer"]),
             "amount": r["amount"]} for r in rev
        ]

    # ── Q2 2026 month filter (order files only — no revenue needed here) ──────
    Q2_MONTHS = {"Apr 2026", "May 2026", "Jun 2026"}
    inc_q2 = [r for r in inc if r.get("month") in Q2_MONTHS]
    cl_q2  = [r for r in cl  if r.get("month") in Q2_MONTHS]

    Q2_TARGET = 20000000   # ₹2 Cr

    # ── Pipeline summary ──────────────────────────────────────────────────────
    # Source sheet has the typo "Exisitng Customer" — match both spellings,
    # and also drop garbage rows where probability is outside [0,1] (the script
    # picks up some misaligned rows from the Plan sheet otherwise).
    pipeline = [p for p in qt.get("pipeline", []) if 0 <= p.get("probability", 0) <= 1]
    def _is_existing(t): return "Exist" in t or "Exisit" in t or "Customer" in t
    existing_pipe = [p for p in pipeline if _is_existing(p.get("type",""))]
    prospect_pipe = [p for p in pipeline if "Prospect" in p.get("type","")]

    # ── Insurer-wise Q2 orders ────────────────────────────────────────────────
    def group_by_insurer(records, value_key):
        """Aggregate by canonical insurer name so variant spellings are merged."""
        out = {}
        for r in records:
            ins = _canonical_insurer_name(r["insurer"])
            out[ins] = out.get(ins, 0) + r[value_key]
        return out

    inc_by_insurer = group_by_insurer(inc_q2, "count")
    cl_by_insurer  = group_by_insurer(cl_q2,  "count")
    # rev_by_insurer is already set from billing files above (or fallback path)

    # ── Orders broken down BY MONTH (for the new month-filter tabs + trend) ──
    # Structure: { "Apr 2026": {"incoming": {insurer: n, …}, "closed": {…}}, … }
    def by_month_by_insurer(records, value_key):
        out = {}
        for r in records:
            m = (r.get("month") or "").strip()
            if not m:
                continue
            out.setdefault(m, {})
            ins = r["insurer"]
            out[m][ins] = out[m].get(ins, 0) + r[value_key]
        return out

    inc_monthly = by_month_by_insurer(inc, "count")
    cl_monthly  = by_month_by_insurer(cl,  "count")

    # All months present in either incoming or closed
    all_order_months = sorted(set(inc_monthly) | set(cl_monthly),
                              key=lambda m: _month_sort_key(m))

    orders_monthly = {}
    for m in all_order_months:
        orders_monthly[m] = {
            "incoming": inc_monthly.get(m, {}),
            "closed":   cl_monthly.get(m, {}),
        }

    # Flat list for the monthly-trend bar/line chart — last 15 months
    orders_monthly_totals = []
    for m in all_order_months:
        orders_monthly_totals.append({
            "month":    m,
            "incoming": sum(inc_monthly.get(m, {}).values()),
            "closed":   sum(cl_monthly.get(m, {}).values()),
        })

    # All known insurers — union of order files + billing data so no insurer is missed
    all_insurers = sorted(set(list(inc_by_insurer) + list(cl_by_insurer) + list(rev_by_insurer)))
    insurer_table = []
    for ins in all_insurers:
        i = inc_by_insurer.get(ins, 0)
        c = cl_by_insurer.get(ins, 0)
        r = rev_by_insurer.get(ins, 0)
        insurer_table.append({
            "insurer": ins,
            "incoming": i,
            "closed": c,
            "conversion": round(c/i, 4) if i > 0 else 0,
            "revenue": r,
        })

    # ── Monthly trend (last 15 months, dynamic) ──────────────────────────────
    from calendar import month_abbr
    _now = datetime.now(timezone.utc)
    _months = []
    for i in range(14, -1, -1):
        _m = _now.month - i
        _y = _now.year
        while _m <= 0:
            _m += 12
            _y -= 1
        _months.append(f"{month_abbr[_m]} {_y}")
    # Also include any months from the data not already in the window
    # (keeps historical data if the window shifts)
    billing_order = _months
    trend_labels  = billing_order
    trend_values  = [monthly_billing.get(m, 0) for m in billing_order]

    # ── Apr customer breakdown (from billing Apr data) ───────────────────────
    # Build from Apr billing per-insurer; cross-reference order counts for
    # conversion rate.  Falls back to insurer_table when no Apr billing data.
    apr_customers_billing = [
        {
            "insurer":      ins,
            "revenue":      apr_rev_by_ins.get(ins, 0),
            "closed_cases": cl_by_insurer.get(ins, 0),
            "incoming":     inc_by_insurer.get(ins, 0),
            "conversion":   round(cl_by_insurer.get(ins, 0) /
                                  inc_by_insurer.get(ins, 1), 4)
                            if inc_by_insurer.get(ins, 0) > 0 else 0,
            "avg_order":    0,
        }
        for ins in sorted(
            set(list(apr_rev_by_ins)) | set(list(cl_by_insurer))
        )
        if apr_rev_by_ins.get(ins, 0) > 0
    ]
    # Sort descending by Apr revenue
    apr_customers_billing.sort(key=lambda r: r["revenue"], reverse=True)
    apr_customers = apr_customers_billing or [
        {"insurer": r["insurer"], "revenue": r["revenue"],
         "closed_cases": r["closed"], "incoming": r["incoming"],
         "conversion": r["conversion"], "avg_order": 0}
        for r in insurer_table if r["revenue"] > 0
    ]

    # ── CSV-based per-insurer revenue (Gmail Revenue_Generated_Insurer_Wise.csv) ──
    # Read separately so we can show the billing-vs-CSV gap in the dashboard.
    import datetime as _dt
    _rev_csv = read_revenue_file(REVENUE_FILE)
    _may_csv_by_ins: dict = {}
    _apr_csv_by_ins: dict = {}
    for _r in _rev_csv:
        _m   = (_r.get("month") or "").strip()
        _ins = _canonical_insurer_name(_r.get("insurer", ""))
        _amt = _r.get("amount", 0) or 0
        if _m == "May 2026":
            _may_csv_by_ins[_ins] = _may_csv_by_ins.get(_ins, 0) + _amt
        elif _m == "Apr 2026":
            _apr_csv_by_ins[_ins] = _apr_csv_by_ins.get(_ins, 0) + _amt
    _may_csv_total = round(sum(_may_csv_by_ins.values()), 2)
    _apr_csv_total = round(sum(_apr_csv_by_ins.values()), 2)
    _rev_csv_mtime_iso = (
        _dt.datetime.fromtimestamp(REVENUE_FILE.stat().st_mtime, tz=_dt.timezone.utc).isoformat()
        if REVENUE_FILE.exists() else None
    )

    # NOTE: We intentionally do NOT override may_rev/apr_rev with CSV totals.
    # The Revenue_Generated_Insurer_Wise.csv is a periodic snapshot that can
    # lag the billing XLS by days (e.g. CSV from May 18 vs XLS from May 30).
    # The XLS transaction-level data is always the authoritative source.
    # The CSV totals are stored separately so the dashboard can show the gap.
    _billing_mtime = max(
        (f.stat().st_mtime if f.exists() else 0)
        for f in [DETAILED_BILLING_FILE1, DETAILED_BILLING_FILE2, BILLING_FILE]
    )
    _csv_mtime = REVENUE_FILE.stat().st_mtime if REVENUE_FILE.exists() else 0
    log.info(f"  Revenue source: XLS billing (Apr ₹{apr_rev:,.0f}, May ₹{may_rev:,.0f}, Jun ₹{jun_rev:,.0f})")
    log.info(f"  CSV snapshot:   Apr ₹{_apr_csv_total:,.0f}, May ₹{_may_csv_total:,.0f} "
             f"(CSV {_dt.datetime.fromtimestamp(_csv_mtime).strftime('%d %b') if _csv_mtime else 'N/A'}, "
             f"XLS {_dt.datetime.fromtimestamp(_billing_mtime).strftime('%d %b') if _billing_mtime else 'N/A'})")

    return {
        "meta": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "q2_target": Q2_TARGET,
            "as_of": datetime.now(timezone.utc).strftime("%-d %b %Y"),
            "revenue_source": "Daily_Insurer_Billing.xls + Daily_Insurer_Billing_2.xls",
        },
        "kpis": {
            "q2_target":   Q2_TARGET,
            "apr_revenue": apr_rev,
            "may_revenue": may_rev,
            "jun_revenue": jun_rev,
            "q2_achieved": q2_total,
            "gap":         Q2_TARGET - q2_total,
            "achievement_pct": round(q2_total / Q2_TARGET * 100, 1),
            "baseline_est":    15000000,
            "delta_required":  5000000,
            "delta_weighted":  round(qt.get("may_pipeline_total",0) + qt.get("june_pipeline_total",0)),
        },
        "monthly_trend": {
            "labels": trend_labels,
            "values": trend_values,
            "by_insurer": {
                m: bk["monthly_by_insurer"].get(m, {})
                for m in trend_labels
            } if bk and bk.get("monthly_by_insurer") else {},
        },
        # Q2 monthly targets — Apr ₹60L + May ₹65L + Jun ₹75L = ₹2 Cr
        "month_wise": [
            {"month":"April 2026",  "target":6000000,  "achieved":apr_rev, "pct": round(apr_rev/6000000*100,1)},
            {"month":"May 2026",    "target":6500000,  "achieved":may_rev, "pct": round(may_rev/6500000*100,1)},
            {"month":"June 2026",   "target":7500000,  "achieved":jun_rev, "pct": round(jun_rev/7500000*100,1) if jun_rev else 0},
        ],
        "apr_customers": apr_customers,
        "insurer_table": insurer_table,
        # CSV-based revenue per insurer (Gmail Revenue_Generated_Insurer_Wise.csv)
        "may_revenue_csv_by_insurer": {k: round(v,2) for k,v in _may_csv_by_ins.items()},
        "apr_revenue_csv_by_insurer": {k: round(v,2) for k,v in _apr_csv_by_ins.items()},
        "may_revenue_csv_total": _may_csv_total,
        "apr_revenue_csv_total": _apr_csv_total,
        "revenue_csv_mtime": _rev_csv_mtime_iso,
        # CSV-based revenue (Gmail Revenue_Generated_Insurer_Wise.csv)
        # Used by the dashboard to show billing-vs-CSV gap.
        "may_revenue_csv_by_insurer": {k: round(v,2) for k,v in _may_csv_by_ins.items()},
        "apr_revenue_csv_by_insurer": {k: round(v,2) for k,v in _apr_csv_by_ins.items()},
        "may_revenue_csv_total": _may_csv_total,
        "apr_revenue_csv_total": _apr_csv_total,
        "revenue_csv_mtime": _rev_csv_mtime_iso,
        "pipeline": pipeline,
        "existing_pipeline_total": round(sum(p["weighted"] for p in existing_pipe)),
        "prospect_pipeline_total": round(sum(p["weighted"] for p in prospect_pipe)),
        "pipeline_by_type": {
            "existing_weighted": round(sum(p["weighted"] for p in existing_pipe)),
            "prospect_weighted": round(sum(p["weighted"] for p in prospect_pipe)),
            "existing_delta":    round(sum(p["delta"] for p in existing_pipe)),
            "prospect_delta":    round(sum(p["delta"] for p in prospect_pipe)),
        },
        "q2_orders": {
            "incoming_by_insurer": inc_by_insurer,
            "closed_by_insurer":   cl_by_insurer,
        },
        # Per-month breakdown — drives the new month-filter tabs in the dashboard.
        "orders_monthly":         orders_monthly,
        "orders_monthly_totals":  orders_monthly_totals[-15:],   # keep last 15 months for chart
        # Required-to-hit-target numbers from the Plan sheet's May/June sections
        "monthly_targets":        qt.get("monthly_targets", {}),
        # Sheet2-driven action items consumed by renderActions()
        "key_actions": qt.get("key_actions", []),
        # Pipeline totals by month (used by the dashboard tabs)
        "pipeline_by_month": {
            "may":   round(sum(p["weighted"] for p in pipeline if p.get("month") == "May")),
            "june":  round(sum(p["weighted"] for p in pipeline if p.get("month") == "June")),
            "q2":    round(sum(p["weighted"] for p in pipeline if p.get("month") == "Q2")),
        },
        "_plan_has_month_col": qt.get("plan_has_month_col", False),
        # ── YoY comparison — sourced from detailed billing files ─────────────
        "yoy_comparison": _compute_yoy(yoy_src),

        # ── All-time billing revenue per canonical insurer ───────────────────
        # Sourced from Daily_Insurer_Billing*.xls (single source of truth).
        # Used by the Insurer Details tab "Overall Revenue Since Inception" column.
        "billing_lifetime_by_insurer": lt_by_ins,

        # ── Detailed category-level billing analysis ──────────────────────────
        # Already computed above from the same detailed_records — no double-read.
        "billing_analysis": billing_analysis_data,
    }


def _billing_lifetime(billing_records: list) -> dict:
    """Sum all-time actual billed revenue per canonical insurer name."""
    from collections import defaultdict
    totals: dict = defaultdict(float)
    for r in billing_records:
        ins = r.get("insurer", "")
        if ins:
            totals[ins] += r.get("amount", 0)
    return {k: round(v, 2) for k, v in totals.items()}


# ─── Detailed billing analysis (category-level breakdown) ─────────────────────

# Noise patterns to exclude from category billing analysis
_BILLING_NOISE = {
    'Cardiotrack Corporate Services', 'Cardiotrack Home Services',
    'India First Life Insurance', 'Health Assure Tata AIA',
    'Shriram Life Insurance', 'Aditya Birla Health Insurance',
    'Pramerica Life Insurance VMER', 'MD India Tata AIA', 'MD India Max Life',
}

def read_detailed_billing_files() -> list:
    """Load Daily_Insurer_Billing.xls and Daily_Insurer_Billing_2.xls.

    Both files share identical columns (ID, Insurer Name, …, Total Amount to be
    Billed, …).  Returns a list of flat dicts with canonical insurer name and all
    six revenue category amounts.

    Column indices (0-based):
        1  Insurer Name
        9  Order Closure Date
       12  Billing Rate for Core Package   → pkg
       13  Insurer Approved Amount         → anc
       14  Videography + Digital MER Rate  → video
       15  Service Charges                 → svc
       17  Home Visit Charges              → hv
       18  Interpretation Charges          → interp
       19  Total Amount to be Billed       → total
    """
    import io as _io
    from collections import defaultdict as _dd

    records: list = []
    for path in (DETAILED_BILLING_FILE1, DETAILED_BILLING_FILE2):
        if not path.exists():
            log.warning(f"Detailed billing file not found: {path.name}")
            continue
        try:
            with open(path, "rb") as fh:
                raw = _io.BytesIO(fh.read())
            wb = openpyxl.load_workbook(raw, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            log.warning(f"Cannot open {path.name}: {e}")
            continue

        if len(all_rows) < 2:
            continue

        for row in all_rows[1:]:
            if not row or len(row) < 20:
                continue
            ins_raw = row[1]
            date_val = row[9]
            if ins_raw is None or date_val is None:
                continue

            ins_str = str(ins_raw).strip()
            if (ins_str in _BILLING_NOISE or ins_str.startswith('Test')
                    or 'Purge' in ins_str):
                continue

            ins = _canonical_insurer_name(ins_str)
            if not ins:
                continue

            if not isinstance(date_val, datetime):
                continue

            def _sf(v):
                try: return float(v) if v not in (None, '') else 0.0
                except: return 0.0

            records.append({
                "insurer":   ins,
                "date":      date_val,
                "month":     date_val.strftime("%b %Y"),
                "pkg":       _sf(row[12]),
                "anc":       _sf(row[13]),
                "video":     _sf(row[14]),
                "svc":       _sf(row[15]),
                "hv":        _sf(row[17]),
                "interp":    _sf(row[18]),
                "total":     _sf(row[19]),
            })

    log.info(f"  Detailed billing: {len(records)} rows from "
             f"{sum(1 for p in (DETAILED_BILLING_FILE1,DETAILED_BILLING_FILE2) if p.exists())} files")
    return records


def build_billing_analysis(records: list) -> dict:
    """Aggregate detailed billing records into the billing_analysis JSON block.

    Periods:
        all   → all available history
        fy    → FY 2025-26  (Apr 2025 – present)
        fy_prev → FY 2024-25 (Apr 2024 – Mar 2025)
        q2    → Q2 2026 (Apr 2026 – present)
    """
    from collections import defaultdict as _dd

    if not records:
        return {}

    FY_START   = datetime(2025, 4,  1)
    FY_PREV_S  = datetime(2024, 4,  1)
    FY_PREV_E  = datetime(2025, 3, 31, 23, 59, 59)
    Q2_START   = datetime(2026, 4,  1)
    CATS       = ["pkg", "anc", "video", "svc", "hv", "interp", "total"]

    def _empty():
        return {c: 0.0 for c in CATS + ["cases"]}

    all_s  = _dd(_empty)
    fy_s   = _dd(_empty)
    fyp_s  = _dd(_empty)
    q2_s   = _dd(_empty)
    monthly: dict = _dd(lambda: _dd(float))

    date_min = date_max = None

    for r in records:
        ins = r["insurer"]
        d   = r["date"]

        for bucket in ([all_s] +
                       ([fy_s]  if d >= FY_START  else []) +
                       ([fyp_s] if FY_PREV_S <= d <= FY_PREV_E else []) +
                       ([q2_s]  if d >= Q2_START  else [])):
            for c in CATS:
                bucket[ins][c] += r[c]
            bucket[ins]["cases"] += 1

        monthly[r["month"]][ins] += r["total"]

        if date_min is None or d < date_min: date_min = d
        if date_max is None or d > date_max: date_max = d

    # Sort insurers by all-time total desc
    ins_order = sorted(all_s, key=lambda i: -all_s[i]["total"])

    def _row(bucket, ins):
        b = bucket.get(ins, {})
        return {c: round(b.get(c, 0), 2) for c in CATS + ["cases"]}

    by_insurer = []
    for rank, ins in enumerate(ins_order, 1):
        a   = all_s[ins]
        tot = a["total"]
        grand_all = sum(all_s[i]["total"] for i in ins_order)
        fy_tot   = fy_s.get(ins, {}).get("total", 0)
        fyp_tot  = fyp_s.get(ins, {}).get("total", 0)
        yoy_pct  = round((fy_tot - fyp_tot) / fyp_tot * 100, 1) if fyp_tot else None
        by_insurer.append({
            "insurer":      ins,
            "rank":         rank,
            "all":          _row(all_s,  ins),
            "fy":           _row(fy_s,   ins),
            "fy_prev":      _row(fyp_s,  ins),
            "q2":           _row(q2_s,   ins),
            "share_pct":    round(tot / grand_all * 100, 1) if grand_all else 0,
            "yoy_pct":      yoy_pct,
            "active_fy":    fy_tot > 0,
            "active_q2":    q2_s.get(ins, {}).get("total", 0) > 0,
        })

    # Monthly totals (all insurers combined) — last 24 months
    sorted_months = sorted(monthly.keys(),
                           key=lambda m: datetime.strptime(m, "%b %Y"))
    monthly_totals = [
        {"month": m, "total": round(sum(monthly[m].values()), 2)}
        for m in sorted_months[-24:]
    ]

    # Grand totals
    grand_all  = sum(all_s[i]["total"]  for i in ins_order)
    grand_fy   = sum(fy_s.get(i,{}).get("total",0)  for i in ins_order)
    grand_fyp  = sum(fyp_s.get(i,{}).get("total",0) for i in ins_order)
    grand_q2   = sum(q2_s.get(i,{}).get("total",0)  for i in ins_order)
    grand_cases= sum(int(all_s[i]["cases"]) for i in ins_order)

    # Category totals (all-time, for mix chart)
    cat_totals = {c: round(sum(all_s[i][c] for i in ins_order), 2) for c in CATS}

    return {
        "generated_at":  datetime.now(timezone.utc).isoformat(),
        "total_rows":    len(records),
        "date_from":     date_min.strftime("%b %Y") if date_min else None,
        "date_to":       date_max.strftime("%b %Y") if date_max else None,
        "grand_totals":  {
            "all_time": round(grand_all, 2),
            "fy_2526":  round(grand_fy, 2),
            "fy_prev":  round(grand_fyp, 2),
            "q2_2026":  round(grand_q2, 2),
            "cases":    grand_cases,
            "yoy_pct":  round((grand_fy - grand_fyp) / grand_fyp * 100, 1) if grand_fyp else None,
        },
        "cat_totals":    cat_totals,
        "by_insurer":    by_insurer,
        "monthly_totals": monthly_totals,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 3.  DASHBOARD HTML BUILDER
# ══════════════════════════════════════════════════════════════════════════════

DATA_PLACEHOLDER = "__DASHBOARD_DATA__"
# Anything we drop in to replace the placeholder so the file stays templatable.
def _embedded_data_line(data: dict) -> str:
    return json.dumps(data, ensure_ascii=False)


def rebuild_html(data: dict):
    """Inject fresh JSON into the dashboard HTML.

    The template is `Automation/dashboard_template.html` if it exists; otherwise
    we re-use the existing output HTML (which carries the placeholder on first
    run). After a successful rebuild we always write back to the output path
    *and* refresh the template, so the placeholder stays available for next time.
    """
    html_path     = OUTPUT / "Cardiotrack_Dashboard.html"
    template_path = AUTO   / "dashboard_template.html"

    if template_path.exists():
        html = template_path.read_text(encoding="utf-8")
    elif html_path.exists() and DATA_PLACEHOLDER in html_path.read_text(encoding="utf-8"):
        # Bootstrap: pre-rebuild output still has the placeholder — promote it
        # to a template so we have something to inject into next time.
        html = html_path.read_text(encoding="utf-8")
        template_path.write_text(html, encoding="utf-8")
        log.info(f"  ↪ Bootstrapped template: {template_path.name}")
    else:
        log.warning(
            "No template found and current HTML has no placeholder — skipping HTML rebuild. "
            "Restore Automation/dashboard_template.html to fix."
        )
        return

    html = html.replace(DATA_PLACEHOLDER, _embedded_data_line(data))
    html_path.write_text(html, encoding="utf-8")
    log.info(f"  ✓ HTML dashboard rebuilt → {html_path.name}")

    # Copy the Account Managers dashboard sibling into Data/Output/ so the
    # main dashboard's iframe (src="Account_Managers_Dashboard.html") can
    # resolve. The AM dashboard is itself a hand-maintained static file;
    # we just keep the deployed copy in sync with whatever is at the repo root.
    am_source = BASE / "Account_Managers_Dashboard.html"
    if am_source.exists():
        am_dest = OUTPUT / "Account_Managers_Dashboard.html"
        am_dest.write_bytes(am_source.read_bytes())
        log.info(f"  ✓ Account Managers dashboard copied → {am_dest.name}")
    else:
        log.debug(f"  (no AM dashboard at {am_source} — skipping copy)")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def run_sync(skip_gmail: bool = False, require_gmail: bool = False,
             force_gmail: bool = False) -> dict:
    """End-to-end: pull from Gmail (optional), rebuild JSON, rebuild HTML.

    Args:
        skip_gmail:    If True, don't talk to Gmail at all (just rebuild from
                       whatever CSV/Excel files are already on disk).
        require_gmail: If True, a Gmail failure (auth, packages, no matching
                       emails) raises RuntimeError instead of silently continuing.
                       The dashboard "Sync Gmail" button sets this; the
                       background daily job and "Refresh" button do not.

    Returns the freshly built dashboard dict with a `gmail_status` field
    embedded so callers can surface what happened.
    """
    gmail_status = None
    if not skip_gmail:
        log.info(f"Step 1: Gmail sync{' (force mode)' if force_gmail else ''}…")
        gmail_status = sync_gmail(force=force_gmail)
        if not gmail_status.get("ok"):
            log.error(f"  Gmail step failed: {gmail_status.get('error')}")
            if require_gmail:
                raise RuntimeError(gmail_status.get("error") or "Gmail sync failed")
    else:
        log.info("Step 1: Gmail sync skipped")

    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl not installed — run: pip3 install openpyxl")

    log.info("Step 2: Processing Excel data…")
    data = build_dashboard_data()

    # Embed the Gmail status in the dashboard data so the UI can surface it.
    if gmail_status is not None:
        data.setdefault("meta", {})["gmail_status"] = gmail_status

    # Embed the public Refresh-button Worker URL.
    # Priority: env var PUBLIC_REFRESH_URL (set via GitHub Actions secret)
    # → local_config.json (local dev) → hardcoded fallback.
    try:
        import os as _os
        _worker_url = (
            _os.environ.get("PUBLIC_REFRESH_URL", "").strip()
            or ""
        )
        if not _worker_url:
            cfg_path = AUTO / "local_config.json"
            if cfg_path.exists():
                cfg = json.loads(cfg_path.read_text())
                _worker_url = (cfg.get("public_refresh_url") or "").strip()
        # Hardcoded fallback — always works even if secret/config is missing
        if not _worker_url or _worker_url.startswith("PASTE_"):
            _worker_url = "https://cardiotrack-sync-trigger.islam-najmul.workers.dev"
        data.setdefault("meta", {})["public_refresh_url"] = _worker_url
    except Exception:
        data.setdefault("meta", {})["public_refresh_url"] =             "https://cardiotrack-sync-trigger.islam-najmul.workers.dev"

    # Also stamp each input file's mtime so the UI can show "Revenue data is
    # from a CSV last refreshed at <time>" — answers the question
    # "is the displayed number actually fresh?"
    meta_files = {}
    for label, path in (("revenue_csv",          REVENUE_FILE),
                        ("incoming_csv",         INCOMING_FILE),
                        ("closed_csv",           CLOSED_FILE),
                        ("billing_csv",          BILLING_FILE),
                        ("quarter_target_xlsx",  QUARTER_TARGET_FILE)):
        if path.exists():
            mt = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
            meta_files[label] = {
                "name":  path.name,
                "mtime": mt.isoformat(),
                "size":  path.stat().st_size,
            }
    data.setdefault("meta", {})["files"] = meta_files

    log.info("Step 3: Writing dashboard_data.json…")
    OUTPUT.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    log.info(f"  ✓ {DATA_JSON.name}")

    log.info("Step 4: Rebuilding HTML dashboard…")
    rebuild_html(data)
    log.info("✅ Done.")
    return data


def diagnose_gmail():
    """Walk every search hit and dump what the script sees — for debugging
    'why isn't this email being picked up?' questions.

    Run:  python3 Automation/cardiotrack_sync.py --gmail-debug
    """
    if not GMAIL_AVAILABLE:
        print("Google API packages not installed.")
        return
    if not CREDS_FILE.exists():
        print(f"credentials.json not found at {CREDS_FILE}")
        return

    try:
        service = gmail_authenticate()
    except Exception as e:
        print(f"Gmail auth failed: {e}")
        return

    sync_data = load_sync_log()
    processed = set(sync_data["processed_message_ids"])
    seen_hashes = sync_data["attachment_hashes"]

    print("=" * 78)
    print(f"GMAIL DIAGNOSTIC — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  processed_message_ids: {len(processed)}")
    print(f"  attachment_hashes:     {len(seen_hashes)}")
    print("=" * 78)

    for report_name, dest_file in EMAIL_MAP.items():
        print(f"\n── REPORT: {report_name} ──")
        print(f"  Destination: {dest_file.name}")
        expected_full = _expected_subject(report_name)
        print(f"  Expected subject contains: {expected_full!r}")

        query = f'subject:"{report_name}" has:attachment filename:csv'
        print(f"  Gmail query: {query}")

        try:
            result = service.users().messages().list(
                userId="me", q=query, maxResults=10
            ).execute()
        except Exception as e:
            print(f"  ✗ Query failed: {e}")
            continue

        msg_refs = result.get("messages", [])
        print(f"  Gmail returned: {len(msg_refs)} message(s)")
        if not msg_refs:
            # Try a looser query to see if the inner phrase exists at all
            loose = f'"{report_name}"'
            loose_result = service.users().messages().list(
                userId="me", q=loose, maxResults=5
            ).execute()
            loose_refs = loose_result.get("messages", [])
            print(f"  Looser query {loose!r} returned: {len(loose_refs)} message(s) — they may be missing the CSV attachment.")
            continue

        for ref in msg_refs:
            msg_id = ref["id"]
            message = service.users().messages().get(
                userId="me", id=msg_id, format="full"
            ).execute()
            subject = _get_header(message, "Subject")
            from_h  = _get_header(message, "From")
            date_h  = _get_header(message, "Date")
            in_processed = "yes" if msg_id in processed else "no"
            subject_ok = _subject_matches(subject, report_name)

            csv_parts = []
            other_parts = []
            for p in _iter_parts(message.get("payload", {})):
                fn = p.get("filename") or ""
                mt = p.get("mimeType") or ""
                has_body = bool((p.get("body") or {}).get("attachmentId") or
                                (p.get("body") or {}).get("data"))
                if fn:
                    if fn.lower().endswith(".csv"):
                        csv_parts.append((fn, mt, has_body))
                    else:
                        other_parts.append((fn, mt))

            verdict = "WILL DOWNLOAD" if (subject_ok and csv_parts and msg_id not in processed) else "skipped"
            mark = "✓" if verdict == "WILL DOWNLOAD" else "•"

            print(f"  {mark} [{msg_id[:14]}…] {date_h}")
            print(f"      From:        {from_h}")
            print(f"      Subject:     {subject!r}   (matches? {subject_ok})")
            print(f"      In processed:{in_processed}")
            print(f"      CSV parts:   {csv_parts or '(none)'}")
            if other_parts:
                print(f"      Other parts: {other_parts}")
            print(f"      → Verdict:   {verdict}")


def reset_dedup():
    """Clear the dedup state so the next sync re-evaluates every message.
    Use this when you want to recover from a poisoned blacklist."""
    if SYNC_LOG.exists():
        SYNC_LOG.unlink()
        print(f"✓ Removed {SYNC_LOG} — next sync will re-evaluate every message.")
    else:
        print(f"No sync log found at {SYNC_LOG} — nothing to clear.")


def main():
    if "--auth" in sys.argv or "--reauth" in sys.argv:
        do_auth()
        return
    if "--gmail-debug" in sys.argv:
        diagnose_gmail()
        return
    if "--reset-dedup" in sys.argv:
        reset_dedup()
        return

    rebuild_only = "--rebuild" in sys.argv
    force_mode   = "--force" in sys.argv
    log.info("=" * 60)
    mode_label = "rebuild only" if rebuild_only else ("force sync" if force_mode else "full sync")
    log.info(f"Cardiotrack Sync  ({mode_label})")
    log.info("=" * 60)
    run_sync(skip_gmail=rebuild_only, force_gmail=force_mode)


if __name__ == "__main__":
    main()
